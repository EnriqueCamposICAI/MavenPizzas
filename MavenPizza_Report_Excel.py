import openpyxl as xl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
from openpyxl.chart import BarChart, Reference, Series, ScatterChart, PieChart


def extract(nombres):
    archivos = []
    for i in range(len(nombres)):
        archivos.append(pd.read_csv('Datasets_pizza_fixed/'+nombres[i]))

    return archivos


def hoja_ejecutiva(ws, archivos):
    '''
    - Tipo de pizzas más vendidad
    - Dinero generado por tipo de pizza
    '''
    archivos[4].index = archivos[4]['pizza_id']
    dict_tipo_pizzas = {}
    dict_tamaño_pizzas = {'S': 0, 'M': 0, 'L': 0, 'XL': 0, 'XXL': 0}
    for i in range(len(archivos[1].index)):
        pedido = archivos[1].loc[i, 'pizza_id']
        pizza = re.split(
            r'_', pedido[::-1], 1)

        precio = archivos[4].loc[pedido, 'price']
        tamaño_pizza = pizza[0][::-1].upper()
        tipo_pizza = pizza[-1][::-1]
        cantidad = archivos[1].loc[i, 'quantity']

        if tipo_pizza in dict_tipo_pizzas:
            dict_tipo_pizzas[tipo_pizza][0] += cantidad
            dict_tipo_pizzas[tipo_pizza][1] += precio*cantidad
        else:
            dict_tipo_pizzas[tipo_pizza] = [cantidad, precio*cantidad]

        dict_tamaño_pizzas[tamaño_pizza] += cantidad

    nombres = {}

    for i in range(archivos[3].shape[0]):
        nombres[archivos[3].loc[i, 'pizza_type_id']
                ] = archivos[3].loc[i, 'name']

    diccionario_pizzas_n = {}

    for key in dict_tipo_pizzas:
        diccionario_pizzas_n[re.sub(
            r'The | Pizza', "", nombres[key])] = dict_tipo_pizzas[key]

    df_tipo = pd.DataFrame.from_dict(
        diccionario_pizzas_n, orient='index', columns=['Cantidad', 'Ingresos'])

    df_tipo.sort_values(by=['Cantidad'], inplace=True, ascending=False)
    for r in dataframe_to_rows(df_tipo, index=True, header=True):
        ws.append(r)
    ws['A1'] = 'Pizza'
    df_tamaño = pd.DataFrame(dict_tamaño_pizzas.items(), columns=[
        'Tamaño', 'Cantidad'])

    for r in dataframe_to_rows(df_tamaño, index=False, header=True):
        ws.append(r)

    ws.move_range("A35:B40", rows=-34, cols=4)
    ws.move_range('A3:C34', rows=-1, cols=0)

    tipo = Table(displayName="Tipo", ref="A1:C33")
    tamano = Table(displayName="Tamaño", ref="E1:F6")

    style = TableStyleInfo(name="TableStyleMedium18", showFirstColumn=True,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)

    tipo.tableStyleInfo = style
    ws.add_table(tipo)

    tamano.tableStyleInfo = style
    ws.add_table(tamano)

    for cell in ws['A'] + ws[1] + ws['E']:
        cell.font = Font(bold=True)

    ws.column_dimensions['A'].width = 30

    # Charts

    chart_cantidad = BarChart()
    chart_cantidad.type = "bar"
    chart_cantidad.style = 10
    chart_cantidad.title = "Cantidad de pizzas vendidas por tipo de pizza"
    chart_cantidad.y_axis.title = 'Cantidad'
    chart_cantidad.x_axis.title = 'Tipo de pizza'

    data = Reference(ws, min_col=2, min_row=1, max_row=33, max_col=2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=33)

    chart_cantidad.add_data(data, titles_from_data=True)
    chart_cantidad.set_categories(cats)
    chart_cantidad.shape = 4
    chart_cantidad.height = 20
    chart_cantidad.width = 25
    ws.add_chart(chart_cantidad, "E9")

    chart_dinero = BarChart()
    chart_dinero.type = "bar"
    chart_dinero.style = 10
    chart_dinero.title = "Cantidad de ingresos por tipo de pizza"
    chart_dinero.y_axis.title = 'Ingresos'
    chart_dinero.x_axis.title = 'Tipo de pizza'

    data = Reference(ws, min_col=3, min_row=1, max_row=33, max_col=3)
    cats = Reference(ws, min_col=1, min_row=2, max_row=33)

    chart_dinero.add_data(data, titles_from_data=True)
    chart_dinero.set_categories(cats)
    chart_dinero.shape = 4
    chart_dinero.height = 20
    chart_dinero.width = 25
    ws.add_chart(chart_dinero, "S9")

    scatter = ScatterChart(scatterStyle='marker')
    scatter.title = "Cantidad de pizzas vendidas por ingresos"
    scatter.style = 13
    scatter.x_axis.title = 'Cantidad'
    scatter.y_axis.title = 'Ingresos'

    xvalues = Reference(ws, min_col=2, min_row=2, max_row=33)
    values = Reference(ws, min_col=3, min_row=2, max_row=33)
    series = Series(values, xvalues, title_from_data=False,
                    title='Tipo de pizza unico')
    series.marker = xl.chart.marker.Marker('x')
    series.graphicalProperties.line.noFill = True
    scatter.series.append(series)

    ws.add_chart(scatter, "E50")

    chart_tamaño = PieChart()
    labels = Reference(ws, min_col=5, min_row=2, max_row=6)
    data = Reference(ws, min_col=6, min_row=1, max_row=6)
    chart_tamaño.add_data(data, titles_from_data=True)
    chart_tamaño.set_categories(labels)
    chart_tamaño.title = "Cantidad de pizzas vendidas por tamaño"

    ws.add_chart(chart_tamaño, "S50")

    return ws


def hoja_ingredientes(ws):
    '''
    - Tabla de ingredientes
    '''
    ingredientes = pd.read_csv(
        "Datasets_pizza_fixed/Prediccion_ingredientes_Final_2016.csv", encoding="utf-8")
    ingredientes.rename(
        columns={'Unnamed: 0': 'Ingrediente'}, inplace=True)

    columnas = ingredientes.columns

    columnas_n = [columnas[0], columnas[-1]]
    for i in range(1, len(columnas)-1):
        columnas_n.append(columnas[i])

    ingredientes = ingredientes[columnas_n]

    for r in dataframe_to_rows(ingredientes, index=False, header=True):
        ws.append(r)

    ingredientes = Table(displayName="Ingredientes", ref="A1:BD67")

    style = TableStyleInfo(name="TableStyleMedium18", showFirstColumn=True,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)

    ingredientes.tableStyleInfo = style
    ws.add_table(ingredientes)

    for cell in ws['B'] + ws[1] + ws['A']:
        cell.font = Font(bold=True)

    for cell in ws[1]:
        letter = cell.column_letter
        ws.column_dimensions[letter].width = 12

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    return ws


def hoja_pedidos(ws, archivos):
    '''
    - Franjas horarias de pedidos
    - Meses de los pedidos
    '''
    dict_horas = {'00': 0, '01': 0, '02': 0, '03': 0, '04': 0, '05': 0, '06': 0, '07': 0, '08': 0, '09': 0, '10': 0,
                  '11': 0, '12': 0, '13': 0, '14': 0, '15': 0, '16': 0, '17': 0, '18': 0, '19': 0, '20': 0, '21': 0, '22': 0, '23': 0}
    dict_meses = {'01': 0, '02': 0, '03': 0, '04': 0, '05': 0,
                  '06': 0, '07': 0, '08': 0, '09': 0, '10': 0, '11': 0, '12': 0}

    for i in range(len(archivos[2].index)):
        hora = archivos[2].loc[i, 'time'].split(':')[0]
        mes = archivos[2].loc[i, 'date'].split('/')[1]

        if hora in dict_horas:
            dict_horas[hora] += 1
        else:
            dict_horas[hora] = 1

        if mes in dict_meses:
            dict_meses[mes] += 1
        else:
            dict_meses[mes] = 1
    lista_meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto',
                   'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    dict_meses_l = dict(zip(lista_meses, dict_meses.values()))

    df_horas = pd.DataFrame(dict_horas.items(), columns=['Hora', 'Cantidad'])
    df_meses = pd.DataFrame(dict_meses_l.items(), columns=['Mes', 'Cantidad'])

    df_horas.astype({'Hora': int, 'Cantidad': int})
    for r in dataframe_to_rows(df_meses, index=False, header=True):
        ws.append(r)

    ws.move_range('A1:B13', rows=0, cols=3)

    for r in dataframe_to_rows(df_horas, index=False, header=True):
        ws.append(r)
    ws.move_range('A14:B39', rows=-13, cols=0)

    meses = Table(displayName="Meses", ref="A1:B25")
    horas = Table(displayName="Horas", ref="D1:E13")

    style = TableStyleInfo(name="TableStyleMedium18", showFirstColumn=True,
                           showLastColumn=False, showRowStripes=False, showColumnStripes=False)

    meses.tableStyleInfo = style
    ws.add_table(meses)

    horas.tableStyleInfo = style
    ws.add_table(horas)

    for cell in ws['A'] + ws[1] + ws['D']:
        cell.font = Font(bold=True)

    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['D'].width = 11

    # Charts

    chart_horas = BarChart()
    chart_horas.type = "col"
    chart_horas.style = 10
    labels = Reference(ws, min_col=1, min_row=2, max_row=25)
    data = Reference(ws, min_col=2, min_row=1, max_row=25)
    chart_horas.add_data(data, titles_from_data=True)
    chart_horas.set_categories(labels)
    chart_horas.title = "Cantidad de pedidos por franja horaria"
    ws.add_chart(chart_horas, "G4")

    chart_meses = BarChart()
    chart_meses.type = "bar"
    chart_meses.style = 10
    labels = Reference(ws, min_col=4, min_row=2, max_row=13)
    data = Reference(ws, min_col=5, min_row=1, max_row=13)
    chart_meses.add_data(data, titles_from_data=True)
    chart_meses.set_categories(labels)
    chart_meses.title = "Cantidad de pedidos por mes"
    chart_meses.height = 8
    ws.add_chart(chart_meses, "G20")

    return ws


def hoja_de_calculo(archivos):
    wb = xl.Workbook()

    ws1 = wb.active

    ws1.title = 'Reporte Ejecutivo'

    ws1 = hoja_ejecutiva(ws1, archivos)

    ws2 = wb.create_sheet(title='Reporte Ingredientes')

    ws2 = hoja_ingredientes(ws2)

    ws3 = wb.create_sheet(title='Reporte Pedidos')

    ws3 = hoja_pedidos(ws3, archivos)

    wb.save('MavenPizza_Report.xlsx')


def main():
    nombres = ['data_dictionary.csv', 'order_details_fixed.csv',
               'orders_fixed.csv', 'pizza_types.csv', 'pizzas.csv']
    archivos = extract(nombres)
    hoja_de_calculo(archivos)
    pass


if __name__ == "__main__":
    main()
    '''
    Hay que hacer 3 hojas
    1. Hoja ejecutiva
    2. Hoja de ingredientes
    3. Hoja de pedidos
    '''
    pass
